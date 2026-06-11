import os
import tempfile
from flask import Flask, request
import requests
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials
import smtplib
from email.mime.text import MIMEText
from email.header import Header

app = Flask(__name__)

BOT_TOKEN = os.environ.get("BOT_TOKEN")
GOOGLE_SHEET_ID = os.environ.get("GOOGLE_SHEET_ID")
GOOGLE_SERVICE_ACCOUNT_EMAIL = os.environ.get("GOOGLE_SERVICE_ACCOUNT_EMAIL")
GOOGLE_PRIVATE_KEY = os.environ.get("GOOGLE_PRIVATE_KEY")
SMTP_HOST = os.environ.get("SMTP_HOST")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "465"))

SMTP_USER = os.environ.get("SMTP_USER")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD")

EMAIL_FROM = os.environ.get("EMAIL_FROM")
EMAIL_FROM_NAME = os.environ.get("EMAIL_FROM_NAME", "ООО Инвиктика")

EMAIL_SUBJECT = os.environ.get(
    "EMAIL_SUBJECT",
    "Информация о дебиторской задолженности"
)

COPY_EMAIL = os.environ.get("COPY_EMAIL")
PENDING_SENDS = {}
PROCESSING_SENDS = set()


@app.route("/")
def index():
    return "Debt bot is running"


def send_message(chat_id, text, reply_markup=None):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    payload = {"chat_id": chat_id, "text": text}

    if reply_markup:
        payload["reply_markup"] = reply_markup

    requests.post(url, json=payload)


def answer_callback(callback_query_id):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/answerCallbackQuery"
    requests.post(url, json={"callback_query_id": callback_query_id})


def get_telegram_file(file_id):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/getFile"
    data = requests.get(url, params={"file_id": file_id}).json()

    if not data.get("ok"):
        raise Exception("Не удалось получить файл из Telegram")

    file_path = data["result"]["file_path"]
    file_url = f"https://api.telegram.org/file/bot{BOT_TOKEN}/{file_path}"

    file_response = requests.get(file_url)
    file_response.raise_for_status()

    return file_response.content


def normalize_client(name):
    return (
        str(name)
        .strip()
        .lower()
        .replace("«", "")
        .replace("»", "")
        .replace('"', "")
        .replace("ооо ", "")
        .replace("ип ", "")
        .replace("  ", " ")
    )


def get_google_sheet():
    private_key = GOOGLE_PRIVATE_KEY.replace("\\n", "\n")

    info = {
        "type": "service_account",
        "client_email": GOOGLE_SERVICE_ACCOUNT_EMAIL,
        "private_key": private_key,
        "token_uri": "https://oauth2.googleapis.com/token",
    }

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    creds = Credentials.from_service_account_info(info, scopes=scopes)
    client = gspread.authorize(creds)

    return client.open_by_key(GOOGLE_SHEET_ID)


def get_clients_base():
    sheet = get_google_sheet()
    ws = sheet.worksheet("БАЗА_КЛИЕНТОВ")
    rows = ws.get_all_records()

    base = {}

    for row in rows:
        client = str(row.get("Клиент", "")).strip()
        email = str(row.get("Почта", "")).strip()

        if client:
            base[normalize_client(client)] = {
                "client": client,
                "email": email,
            }

    return base


def parse_amount(value):
    if value in [None, ""]:
        return 0

    try:
        return float(str(value).replace(" ", "").replace(",", "."))
    except Exception:
        return 0


def parse_days(value):
    if value in [None, ""]:
        return 0

    try:
        return int(float(str(value).replace(" ", "").replace(",", ".")))
    except Exception:
        return 0


def format_date(value):
    if value in [None, ""]:
        return ""

    if hasattr(value, "strftime"):
        return value.strftime("%d.%m.%Y")

    return str(value)


def analyze_excel(file_content):
    clients_base = get_clients_base()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(file_content)
        tmp_path = tmp.name

    try:
        wb = load_workbook(tmp_path, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        required = [
            "Плательщик",
            "Номер счета",
            "Дата счета",
            "Сумма, не закрытая платежными поручениями",
            "Дней до оплаты",
        ]

        missing = [col for col in required if col not in headers]

        if missing:
            return {
                "error": (
                    "В файле не хватает колонок:\n"
                    + "\n".join(missing)
                    + "\n\nПроверь точные названия заголовков."
                )
            }

        payer_idx = headers.index("Плательщик") + 1
        invoice_idx = headers.index("Номер счета") + 1
        date_idx = headers.index("Дата счета") + 1
        amount_idx = headers.index("Сумма, не закрытая платежными поручениями") + 1
        days_idx = headers.index("Дней до оплаты") + 1

        grouped = {}
        total_rows = 0
        total_sum = 0

        for row in range(2, ws.max_row + 1):
            payer = ws.cell(row=row, column=payer_idx).value
            invoice_number = ws.cell(row=row, column=invoice_idx).value
            invoice_date = ws.cell(row=row, column=date_idx).value
            amount = ws.cell(row=row, column=amount_idx).value
            days_to_pay = ws.cell(row=row, column=days_idx).value

            if not payer:
                continue

            amount_num = parse_amount(amount)

            if amount_num <= 0:
                continue

            total_rows += 1
            total_sum += amount_num

            norm_payer = normalize_client(payer)

            if norm_payer not in grouped:
                grouped[norm_payer] = {
                    "client": str(payer).strip(),
                    "total_sum": 0,
                    "invoices": [],
                }

            grouped[norm_payer]["total_sum"] += amount_num
            grouped[norm_payer]["invoices"].append({
                "invoice_number": str(invoice_number or "").strip(),
                "invoice_date": format_date(invoice_date),
                "amount": amount_num,
                "days_overdue": parse_days(days_to_pay),
            })

        ready = []
        no_email = []
        not_found = []

        for norm_client, data in grouped.items():
            base_item = clients_base.get(norm_client)

            if not base_item:
                not_found.append(data)
                continue

            email = base_item.get("email", "")

            if not email:
                no_email.append(data)
                continue

            data["email"] = email
            ready.append(data)

        return {
            "total_clients": len(grouped),
            "total_rows": total_rows,
            "total_sum": total_sum,
            "ready": ready,
            "no_email": no_email,
            "not_found": not_found,
        }

    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


def build_report(result):
    ready = result["ready"]
    no_email = result["no_email"]
    not_found = result["not_found"]

    text = (
        "Файл прочитан ✅\n\n"
        f"Клиентов с задолженностью: {result['total_clients']}\n"
        f"Строк/счетов с задолженностью: {result['total_rows']}\n"
        f"Общая сумма: {result['total_sum']:,.2f} руб.\n\n"
        f"✅ Готово к рассылке: {len(ready)}\n"
        f"⚠️ Клиент найден, но нет почты: {len(no_email)}\n"
        f"❌ Клиент не найден в базе: {len(not_found)}\n"
    ).replace(",", " ")

    if ready:
        text += "\n\nПервые готовые к рассылке:\n"
        for item in ready[:10]:
            text += (
                f"• {item['client']}: {item['total_sum']:,.2f} руб. "
                f"→ {item['email']}\n"
            ).replace(",", " ")

    if no_email:
        text += "\n\n⚠️ Нет почты:\n"
        for item in no_email[:10]:
            text += f"• {item['client']}: {item['total_sum']:,.2f} руб.\n".replace(",", " ")

    if not_found:
        text += "\n\n❌ Нет в базе:\n"
        for item in not_found[:10]:
            text += f"• {item['client']}: {item['total_sum']:,.2f} руб.\n".replace(",", " ")

    if ready:
        text += "\n\nПроверьте список. Если все верно — подтвердите рассылку."
    else:
        text += "\n\nНет клиентов, готовых к рассылке."

    return text


def build_email_preview(item):
    total_sum = f"{item['total_sum']:,.2f}".replace(",", " ")

    invoices_text = ""

    for invoice in item["invoices"]:
        inv_sum = f"{invoice['amount']:,.2f}".replace(",", " ")

        invoices_text += (
            f"Счет №{invoice['invoice_number']} от {invoice['invoice_date']} — "
            f"{inv_sum} руб., просрочен на {invoice['days_overdue']} дней.\n"
        )

    return (
        f"Кому: {item['email']}\n"
        f"Клиент: {item['client']}\n"
        f"Общая сумма: {total_sum} руб.\n\n"
        "Текст письма:\n"
        "Добрый день!\n\n"
        "Ниже направляем информацию о дебиторской задолженности по вашей компании.\n\n"
        f"Общая сумма задолженности составляет: {total_sum} руб.\n\n"
        f"{invoices_text}\n"
        "По всем вопросам вы можете обратиться к своему менеджеру.\n\n"
    )
def format_days_word(days):
    days = abs(int(days))

    if 11 <= days % 100 <= 14:
        return "дней"

    last_digit = days % 10

    if last_digit == 1:
        return "день"
    elif 2 <= last_digit <= 4:
        return "дня"
    else:
        return "дней"
        
def build_email_body(item):
    total_sum = f"{item['total_sum']:,.2f}".replace(",", " ")

    invoices_text = ""

    for invoice in item["invoices"]:
        inv_sum = f"{invoice['amount']:,.2f}".replace(",", " ")

        days = int(invoice["days_overdue"])

        if days < 0:
           days_abs = abs(days)
           status_text = f"просрочен на {days_abs} {format_days_word(days_abs)}"
        elif days > 0:
           status_text = f"до оплаты {days} {format_days_word(days)}"
        else:
           status_text = "срок оплаты сегодня"

        invoices_text += (
            f"Счет №{invoice['invoice_number']} от {invoice['invoice_date']} — "
            f"{inv_sum} руб., {status_text}.\n"
        )

    return (
        "Добрый день!\n\n"
        "Ниже направляем информацию о дебиторской задолженности по вашей компании.\n\n"
        f"Общая сумма задолженности составляет: {total_sum} руб.\n\n"
        f"{invoices_text}\n"
        "По всем вопросам вы можете обратиться к своему менеджеру.\n\n"
        "Спасибо."
    )


def send_email(to_email, subject, body):
    msg = MIMEText(body, "plain", "utf-8")

    msg["Subject"] = Header(subject, "utf-8")
    msg["From"] = f"{EMAIL_FROM_NAME} <{EMAIL_FROM}>"
    msg["To"] = to_email

    recipients = [to_email]

    if COPY_EMAIL:
        msg["Cc"] = COPY_EMAIL
        recipients.append(COPY_EMAIL)

    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=20) as server:
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.sendmail(
            EMAIL_FROM,
            recipients,
            msg.as_string()
        )


def send_mailing(ready):
    sent = 0
    errors = []

    for item in ready:
        try:
            body = build_email_body(item)

            send_email(
                item["email"],
                EMAIL_SUBJECT,
                body
            )

            sent += 1

        except Exception as e:
            errors.append(
                f"{item['client']} → {item['email']} : {e}"
            )

    return sent, errors
from datetime import datetime


def write_mailing_to_sheet(ready):
    sheet = get_google_sheet()
    ws = sheet.worksheet("РАССЫЛКА")

    rows = []

    for item in ready:
        body = build_email_body(item)

        rows.append([
            datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
            item["client"],
            item["email"],
            EMAIL_SUBJECT,
            body,
            "Готово к отправке",
            ""
        ])

    if rows:
        ws.append_rows(rows, value_input_option="USER_ENTERED")

    return len(rows)
    
def trigger_apps_script_send():
    apps_script_url = os.environ.get("APPS_SCRIPT_URL")
    apps_script_secret = os.environ.get("APPS_SCRIPT_SECRET")

    if not apps_script_url or not apps_script_secret:
        raise Exception(
            "Не заполнены APPS_SCRIPT_URL или APPS_SCRIPT_SECRET"
        )

    response = requests.post(
        apps_script_url,
        json={
            "secret": apps_script_secret
        },
        timeout=60
    )

    response.raise_for_status()

    return response.json()

def build_confirm_keyboard():
    return {
        "inline_keyboard": [
            [
                {"text": "✅ Подтвердить рассылку", "callback_data": "confirm_send"},
                {"text": "❌ Отменить", "callback_data": "cancel_send"},
            ]
        ]
    }


@app.route("/webhook", methods=["POST"])
def webhook():
    data = request.get_json()

    if not data:
        return "ok"

    if "callback_query" in data:
        callback = data["callback_query"]
        callback_id = callback["id"]
        chat_id = callback["message"]["chat"]["id"]
        action = callback.get("data")

        answer_callback(callback_id)

        if action == "cancel_send":
                    if action == "cancel_bulk":
            BULK_SENDS.pop(str(chat_id), None)
            send_message(chat_id, "Массовая рассылка отменена ❌")
            return "ok"

                if action == "confirm_bulk":
            bulk = BULK_SENDS.get(str(chat_id))

            if not bulk:
                return "ok"

            clients = bulk.get("clients", [])
            subject = bulk.get("subject")
            body = bulk.get("body")

            prepared_count = write_bulk_mailing_to_sheet(clients, subject, body)

            send_message(
                chat_id,
                "Массовая рассылка подготовлена ✅\n\n"
                f"Добавлено в лист МАССОВАЯ_РАССЫЛКА: {prepared_count}\n\n"
                "Теперь нужно запустить отправку из Apps Script."
            )

            BULK_SENDS.pop(str(chat_id), None)
            return "ok"

        if action == "confirm_send":
            chat_key = str(chat_id)

            if chat_key in PROCESSING_SENDS:
                return "ok"

            pending = PENDING_SENDS.get(chat_key)

            if not pending:
                return "ok"

            ready = pending.get("ready", [])

            if not ready:
                send_message(chat_id, "Нет клиентов, готовых к рассылке.")
                return "ok"

            PROCESSING_SENDS.add(chat_key)

            try:
                send_message(chat_id, "Подготавливаю письма...")

                prepared_count = write_mailing_to_sheet(ready)

                send_message(
                    chat_id,
                    "Письма подготовлены ✅\n\n"
                    f"Добавлено в лист РАССЫЛКА: {prepared_count}\n\n"
                    "Запускаю отправку писем..."
                )

                script_result = trigger_apps_script_send()

                result_text = (
                    "Рассылка завершена ✅\n\n"
                    f"Отправлено писем: {script_result.get('sent', 0)}\n"
                    f"Ошибок: {script_result.get('errors', 0)}"
                )

                send_message(chat_id, result_text)

                PENDING_SENDS.pop(chat_key, None)

            except Exception as e:
                send_message(
                    chat_id,
                    f"Ошибка при автоматической отправке:\n{e}"
                )

            finally:
                PROCESSING_SENDS.discard(chat_key)

            return "ok"
    message = data.get("message", {})
    chat = message.get("chat", {})
    chat_id = chat.get("id")
    text = message.get("text", "")
    document = message.get("document")

    if not chat_id:
        return "ok"

    if text == "/start":
        send_message(chat_id, "Привет! Я бот по дебиторской задолженности.\n\nКоманда: /дебиторка")
        return "ok"

    if text == "/дебиторка":
        send_message(
            chat_id,
            "Загрузите Excel-файл с дебиторской задолженностью.\n\n"
            "Обязательные колонки:\n"
            "Плательщик | Номер счета | Дата счета | "
            "Сумма, не закрытая платежными поручениями | Дней до оплаты\n\n"
            "Почта подтянется из листа БАЗА_КЛИЕНТОВ."
        )
        return "ok"
    if text == "/рассылка":
        BULK_SENDS[str(chat_id)] = {
            "step": "waiting_text"
        }

        send_message(
            chat_id,
            "Отправьте текст письма для массовой рассылки клиентам.\n\n"
            "Первая строка будет темой письма.\n"
            "Остальной текст — телом письма."
        )
        return "ok"
    if document:
        file_name = document.get("file_name", "")

        if not file_name.endswith((".xlsx", ".xls")):
            send_message(chat_id, "Пожалуйста, загрузите Excel-файл .xlsx или .xls")
            return "ok"

        try:
            send_message(chat_id, "Файл получила. Сверяю с базой клиентов...")
            file_content = get_telegram_file(document["file_id"])
            result = analyze_excel(file_content)

            if "error" in result:
                send_message(chat_id, result["error"])
                return "ok"

            PENDING_SENDS[str(chat_id)] = result

            report = build_report(result)
            send_message(chat_id, report, reply_markup=build_confirm_keyboard())

            ready = result.get("ready", [])
            if ready:
                preview_text = "Предпросмотр первого письма:\n\n" + build_email_preview(ready[0])
                send_message(chat_id, preview_text)

        except Exception as e:
            send_message(chat_id, f"Ошибка при обработке файла:\n{e}")

        return "ok"
    bulk = BULK_SENDS.get(str(chat_id))

    if bulk and bulk.get("step") == "waiting_text" and text:
        lines = text.strip().split("\n")
        subject = lines[0].strip()
        body = "\n".join(lines[1:]).strip()

        if not subject or not body:
            send_message(
                chat_id,
                "Не вижу тему или текст письма.\n\n"
                "Формат:\n"
                "Тема письма\n"
                "Текст письма..."
            )
            return "ok"

        clients = get_all_clients_for_bulk()

        BULK_SENDS[str(chat_id)] = {
            "step": "confirm",
            "subject": subject,
            "body": body,
            "clients": clients
        }

        preview = (
            "Предпросмотр массовой рассылки:\n\n"
            f"Тема: {subject}\n"
            f"Получателей: {len(clients)}\n\n"
            "Текст:\n"
            f"{body[:1500]}"
        )

        keyboard = {
            "inline_keyboard": [
                [
                    {"text": "✅ Отправить всем", "callback_data": "confirm_bulk"},
                    {"text": "❌ Отменить", "callback_data": "cancel_bulk"},
                ]
            ]
        }

        send_message(chat_id, preview, reply_markup=keyboard)
        return "ok"
    send_message(chat_id, "Я получил сообщение ✅\n\nПока понимаю команду: /дебиторка")
    return "ok"
    BULK_SENDS = {}

def get_all_clients_for_bulk():
    sheet = get_google_sheet()
    ws = sheet.worksheet("БАЗА_КЛИЕНТОВ")
    rows = ws.get_all_records()

    clients = []

    for row in rows:
        client = str(row.get("Клиент", "")).strip()
        email = str(row.get("Почта", "")).strip()

        if client and email:
            clients.append({
                "client": client,
                "email": email
            })

    return clients


def write_bulk_mailing_to_sheet(clients, subject, body):
    sheet = get_google_sheet()
    ws = sheet.worksheet("МАССОВАЯ_РАССЫЛКА")

    rows = []

    for item in clients:
        rows.append([
            datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
            item["client"],
            item["email"],
            subject,
            body,
            "Готово к отправке",
            ""
        ])

    if rows:
        ws.append_rows(rows, value_input_option="USER_ENTERED")

    return len(rows)
